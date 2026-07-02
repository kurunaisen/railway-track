from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
EXPLICATIONS_DIR = ROOT / "Экспликации"
OUTPUT_PATH = ROOT / "backend" / "app" / "data" / "railway_explications.json"


ROMAN_PATHS = {
    "I": "1",
    "II": "2",
    "III": "3",
    "IV": "4",
    "V": "5",
    "VI": "6",
    "VII": "7",
    "VIII": "8",
    "IX": "9",
    "X": "10",
}


def _cell(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return re.sub(r"\s+", " ", text)


def _station_name(value: Any) -> str | None:
    text = _cell(value)
    if not text:
        return None
    upper = text.upper()
    if not re.search(r"[A-ZА-ЯЁ]", upper):
        return None
    if "ИТОГО" in upper or upper in {"СТАНЦИЯ", "РЕЗУЛЬТАТ"}:
        return None
    blockpost = re.search(r"БЛОК\s*[- ]?ПОСТ\s+(\d+)\s*КМ|БЛОКПОСТ\s+(\d+)\s*КМ", upper)
    if blockpost:
        km = blockpost.group(1) or blockpost.group(2)
        return f"Блокпост {km} км"
    parts = []
    for part in text.split("-"):
        if re.fullmatch(r"[IVX]+", part, re.IGNORECASE):
            parts.append(part.upper())
        elif part.upper().startswith("ГЭС"):
            parts.append(part.upper())
        else:
            parts.append(part.lower().capitalize())
    return "-".join(parts)


def _path_number(value: Any) -> str | None:
    text = _cell(value)
    if not text or text == "-":
        return None
    roman = ROMAN_PATHS.get(text.upper())
    if roman:
        return roman
    if re.fullmatch(r"\d+/\d+", text):
        return text
    match = re.search(r"\b([IVX]+)\b\s*$", text, re.IGNORECASE)
    if match and match.group(1).upper() in ROMAN_PATHS:
        return ROMAN_PATHS[match.group(1).upper()]
    match = re.search(r"\b(\d+)\b", text)
    return match.group(1) if match else text


def _first_existing(patterns: tuple[str, ...]) -> Path:
    for pattern in patterns:
        matches = sorted(EXPLICATIONS_DIR.glob(pattern))
        if matches:
            return matches[0]
    raise FileNotFoundError(f"No workbook found for patterns: {patterns}")


def _dedupe_rows(rows: list[dict[str, Any]], key_fields: tuple[str, ...]) -> list[dict[str, Any]]:
    seen: set[tuple[str, ...]] = set()
    result: list[dict[str, Any]] = []
    for row in rows:
        key = tuple(str(row.get(field) or "") for field in key_fields)
        if key in seen:
            continue
        seen.add(key)
        result.append(row)
    return result


def read_station_paths(path: Path) -> dict[str, list[dict[str, Any]]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["Экспликация стан путей"]
    stations: dict[str, list[dict[str, Any]]] = {}
    for row in ws.iter_rows(min_row=6, values_only=True):
        station = _station_name(row[0] if len(row) > 0 else None)
        number = _path_number(row[3] if len(row) > 3 else None)
        if not station or not number:
            continue
        item = {
            "number": number,
            "purpose": _cell(row[1] if len(row) > 1 else None) or None,
            "park": _cell(row[2] if len(row) > 2 else None) or None,
            "boundary_from": _cell(row[4] if len(row) > 4 else None) or None,
            "boundary_via": _cell(row[5] if len(row) > 5 else None) or None,
            "boundary_to": _cell(row[6] if len(row) > 6 else None) or None,
            "length_m": _cell(row[7] if len(row) > 7 else None) or None,
            "speed": _cell(row[34] if len(row) > 34 else None) or None,
        }
        stations.setdefault(station, []).append(item)
    return {
        station: _dedupe_rows(rows, ("number", "purpose", "boundary_from", "boundary_to"))
        for station, rows in sorted(stations.items())
    }


def read_switches(path: Path) -> dict[str, list[dict[str, Any]]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["table"]
    stations: dict[str, list[dict[str, Any]]] = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        # In this SAP export, visible data starts at Excel column B.
        station = _station_name(row[1] if len(row) > 1 else None)
        number = _path_number(row[3] if len(row) > 3 else None)
        if not station or not number:
            continue
        item = {
            "number": number,
            "path": _path_number(row[6] if len(row) > 6 else None),
            "path_name": _cell(row[6] if len(row) > 6 else None) or None,
            "purpose": _cell(row[2] if len(row) > 2 else None) or None,
            "park": _cell(row[5] if len(row) > 5 else None) or None,
            "switch_type": _cell(row[8] if len(row) > 8 else None) or None,
            "rail_type": _cell(row[16] if len(row) > 16 else None) or None,
            "frog_mark": _cell(row[17] if len(row) > 17 else None) or None,
            "passenger_speed": _cell(row[18] if len(row) > 18 else None) or None,
            "freight_speed": _cell(row[19] if len(row) > 19 else None) or None,
        }
        stations.setdefault(station, []).append(item)
    return {
        station: _dedupe_rows(rows, ("number", "path", "path_name"))
        for station, rows in sorted(stations.items())
    }


def main() -> None:
    station_paths_file = _first_existing(("Путь*.xlsm", "Экспликация. 2*.xlsm"))
    switches_file = _first_existing(("Стрелочный перевод*.xlsm", "Экспликация стрелочных переводов*.xlsm"))
    station_paths = read_station_paths(station_paths_file)
    switches = read_switches(switches_file)
    station_names = sorted({*station_paths.keys(), *switches.keys()})
    payload = {
        "version": 1,
        "sources": {
            "station_paths": station_paths_file.name,
            "switches": switches_file.name,
        },
        "station_names": station_names,
        "station_paths": station_paths,
        "switches": switches,
    }
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_PATH.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )
    print(
        f"Wrote {OUTPUT_PATH.relative_to(ROOT)}: "
        f"{len(station_names)} stations, "
        f"{sum(len(v) for v in station_paths.values())} paths, "
        f"{sum(len(v) for v in switches.values())} switches"
    )


if __name__ == "__main__":
    main()

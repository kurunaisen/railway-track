"""Excel только из RailwayRow[] — без догадок и извлечений."""

from __future__ import annotations

from io import BytesIO

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.services.railway.types import RailwayRow

DASH = "—"
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)

FORM_COLUMNS: tuple[str, ...] = (
    "Nп/п",
    "Местонахождение (перегон, станция)",
    "№ пути, стрелочного перевода",
    "Привязка (км,пк,м)",
    "Выявленная неисправность",
    "Ограничение скорости",
    "Примечание",
)


def format_asset_for_cell(row: RailwayRow) -> str:
    if row.asset_kind == "switch" and row.asset_number:
        return f"стр. п. {row.asset_number}"
    if row.asset_kind == "track" and row.asset_number:
        return row.asset_number
    return DASH


def format_speed_for_cell(speed: int | None) -> str:
    if speed is None:
        return DASH
    return f"{speed} км/ч"


def railway_rows_to_display_dicts(
    rows: list[RailwayRow],
    *,
    include_source_text: bool = False,
) -> list[dict[str, str]]:
    out: list[dict[str, str]] = []
    for index, row in enumerate(rows, start=1):
        item = {
            "Nп/п": str(index),
            "Местонахождение (перегон, станция)": row.location or DASH,
            "№ пути, стрелочного перевода": format_asset_for_cell(row),
            "Привязка (км,пк,м)": row.reference or DASH,
            "Выявленная неисправность": row.defect or DASH,
            "Ограничение скорости": format_speed_for_cell(row.speed_limit),
            "Примечание": row.note or DASH,
        }
        if include_source_text:
            item["Исходный текст"] = row.source_text or DASH
        out.append(item)
    return out


def export_railway_xlsx(
    rows: list[RailwayRow],
    *,
    include_source_text: bool = False,
    sheet_name: str = "Таблица",
) -> BytesIO:
    display = railway_rows_to_display_dicts(rows, include_source_text=include_source_text)
    columns = list(FORM_COLUMNS)
    if include_source_text:
        columns.append("Исходный текст")

    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        pd.DataFrame(display, columns=columns).to_excel(writer, sheet_name=sheet_name, index=False)
        ws = writer.sheets[sheet_name]
        for col_idx, _ in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.column_dimensions[get_column_letter(col_idx)].width = 22

    buffer.seek(0)
    return buffer

from io import BytesIO
from datetime import datetime

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.models import AudioFile, Export, ProcessingJob
from app.services.evidence_only import is_evidence_only, speed_from_segment
from app.services.inspection_form import build_form_rows, format_object_kind, format_path, format_switch
from app.services.inspection_repository import load_flat_rows, load_latest_done_job
from app.services.session_adapter import audio_file_to_session_out
from app.services.transcript_crypto import decrypt_transcript_text
from app.services.wide_table import build_wide_rows

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def export_session_to_excel(db: Session, session_id: int) -> BytesIO:
    """session_id = audio_files.id."""
    return export_sessions_batch_to_excel(db, [session_id])


def export_sessions_batch_to_excel(db: Session, session_ids: list[int]) -> BytesIO:
    if not session_ids:
        raise ValueError("Не указаны сессии для экспорта")

    all_records = []
    transcript_frames = []
    all_errors: list[dict] = []
    all_terms: list[dict] = []
    export_job_id: int | None = None

    for session_id in session_ids:
        audio = db.query(AudioFile).filter(AudioFile.id == session_id).first()
        if not audio:
            continue
        job = load_latest_done_job(db, session_id)
        if not job:
            continue
        export_job_id = export_job_id or job.id
        records = load_flat_rows(job, session_id)
        all_records.extend(records)
        transcript_frames.append(_sheet_raw_transcripts(session_id, audio, job))
        session_view = audio_file_to_session_out(db, audio)
        for err in session_view.parse_errors:
            all_errors.append({**err, "session_id": session_id})
        for term in session_view.unknown_terms:
            all_terms.append({**term, "session_id": session_id})

    if not all_records:
        raise ValueError("Нет обработанных данных для экспорта")

    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        _sheet_records_form(all_records).to_excel(writer, sheet_name="Таблица", index=False)
        pd.concat(transcript_frames, ignore_index=True).to_excel(
            writer, sheet_name="raw_transcripts", index=False
        ) if transcript_frames else pd.DataFrame().to_excel(
            writer, sheet_name="raw_transcripts", index=False
        )
        _sheet_records_long(all_records).to_excel(writer, sheet_name="records_long", index=False)
        _sheet_records_wide(all_records).to_excel(writer, sheet_name="records_wide", index=False)
        pd.DataFrame(
            [
                {
                    "session_id": e.get("session_id"),
                    "row": e.get("row"),
                    "field": e.get("field"),
                    "error": e.get("error") or e.get("message"),
                    "text_fragment": e.get("text"),
                    "severity": e.get("severity"),
                }
                for e in all_errors
            ]
        ).to_excel(writer, sheet_name="errors", index=False)
        pd.DataFrame(
            [
                {
                    "session_id": t.get("session_id"),
                    "term": t.get("term"),
                    "count": t.get("count"),
                }
                for t in all_terms
            ]
        ).to_excel(writer, sheet_name="unknown_terms", index=False)

        for name in writer.sheets:
            _style_sheet(writer.sheets[name])

    buffer.seek(0)

    if export_job_id:
        label = (
            f"railway_session_{session_ids[0]}.xlsx"
            if len(session_ids) == 1
            else f"railway_batch_{len(session_ids)}_sessions.xlsx"
        )
        db.add(
            Export(
                job_id=export_job_id,
                file_path=label,
                format="xlsx",
                created_at=datetime.utcnow(),
            )
        )
        db.commit()
    return buffer


def _sheet_raw_transcripts(session_id: int, audio: AudioFile, job: ProcessingJob) -> pd.DataFrame:
    rows = []
    transcript = ""
    if job.transcript:
        transcript = decrypt_transcript_text(job.transcript.full_text, job.transcript.text_encrypted) or ""
        for seg in sorted(job.transcript.segments, key=lambda s: s.segment_index):
            rows.append({
                "session_id": session_id,
                "original_name": audio.original_filename,
                "segment_index": seg.segment_index + 1,
                "start_sec": seg.start_sec,
                "end_sec": seg.end_sec,
                "text": seg.text,
                "full_transcript": transcript if seg.segment_index == 0 else "",
            })
    if not rows and transcript:
        rows.append({
            "session_id": session_id,
            "original_name": audio.original_filename,
            "segment_index": 1,
            "start_sec": None,
            "end_sec": None,
            "text": transcript,
            "full_transcript": transcript,
        })
    return pd.DataFrame(rows)


def _sheet_records_long(records) -> pd.DataFrame:
    rows = []
    for idx, rec in enumerate(records, start=1):
        rows.append({
            "№": idx,
            "Дата": rec.record_date,
            "Участок": rec.uchastok,
            "Перегон": rec.peregon,
            "Путь": format_path(rec),
            "Стрелочный перевод": format_switch(rec),
            "Км": rec.km,
            "Пикет": rec.piket,
            "Объект": format_object_kind(rec),
            "Параметр": rec.parameter,
            "Значение": rec.value,
            "Единица": rec.unit,
            "Неисправность/дефект": rec.defect,
            "Комментарий": rec.comment,
            "Ограничение скорости": speed_from_segment(rec) if is_evidence_only() else rec.speed_limit,
            "Таймкод начала": _format_time(rec.segment_start),
            "Таймкод конца": _format_time(rec.segment_end),
            "Спорные поля": ", ".join(rec.disputed_fields),
            "Исходный текст": rec.raw_text,
        })
    return pd.DataFrame(rows)


def _sheet_records_form(records) -> pd.DataFrame:
    _, form_rows = build_form_rows(records)
    if not form_rows:
        return pd.DataFrame(columns=list(build_form_rows([])[0]))
    df = pd.DataFrame(form_rows)
    for col in df.columns:
        df[col] = df[col].apply(lambda v: "—" if v is None or (isinstance(v, str) and not v.strip()) else v)
    return df


def _sheet_records_wide(records) -> pd.DataFrame:
    _, wide_rows = build_wide_rows(records)
    return pd.DataFrame(wide_rows)


def _sheet_errors(session_id: int, errors: list[dict]) -> pd.DataFrame:
    return pd.DataFrame([
        {
            "session_id": session_id,
            "row": e.get("row"),
            "field": e.get("field"),
            "error": e.get("error") or e.get("message"),
            "text_fragment": e.get("text"),
            "severity": e.get("severity"),
        }
        for e in errors
    ])


def _sheet_unknown_terms(session_id: int, terms: list[dict]) -> pd.DataFrame:
    return pd.DataFrame([
        {"session_id": session_id, "term": t.get("term"), "count": t.get("count")}
        for t in terms
    ])


def _style_sheet(ws) -> None:
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 40)


def _format_time(seconds: float | None) -> str:
    if seconds is None:
        return ""
    m, s = divmod(int(seconds), 60)
    h, m = divmod(m, 60)
    if h:
        return f"{h:02d}:{m:02d}:{s:02d}"
    return f"{m:02d}:{s:02d}"
